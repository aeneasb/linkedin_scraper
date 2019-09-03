from linkedin_scraper import Person, actions
from selenium import webdriver
import openpyxl
from openpyxl import Workbook
import time

driver = webdriver.Chrome()

#Log-in to linkedin
email = "Type emaiL"
password = "Type pw"
actions.login(driver, email, password)

# Workbook is created
wb = Workbook()
ws = wb.create_sheet('Longlist Head Ind. Engineering',0)
ws.cell(row=1, column=1, value='name')
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
ws.cell(row=1, column=2, value='Firma / beruflicher Werdegang')
ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
ws.cell(row=1, column=4, value='Aus- und Weiterbildung')

links = ["Python list with linkedin accounts"]

for i,j in enumerate(links):
    person = Person(j, driver=driver,scrape=False)
    time.sleep(1)
    for ss in driver.find_elements_by_class_name("pv-experience-section__see-more"):
        sm = ss.find_element_by_class_name("pv-profile-section__see-more-inline")
        sm.click()

    person.scrape(close_on_complete=False)
    ws.cell(row=i+2, column=1, value=person.name)
    exp_str = ''
    exp_datestr = ''
    edu_str = ''
    edu_datestr = ''

    for exp in person.experiences:
        exp_str += '\n'.join(str(exp).split('\n')[:-1])
        exp_str += '\n'
        exp_datestr += '\n'
        exp_datestr += str(exp).split('\n')[-1]
        exp_datestr += '\n'
    ws.cell(row=i+2, column=3, value=exp_str)
    ws.cell(row=i+2, column=2, value=exp_datestr)

    for edu in person.educations:
        edu_str += str(edu).split('\n')[0]
        edu_str += '\n'
        edu_datestr += str(edu).split('\n')[-1]
        edu_datestr += '\n'
    ws.cell(row=i+2, column=4, value=edu_datestr)
    ws.cell(row=i+2, column=5, value=edu_str)

wb.save('longlist.xlsx')
